import openpyxl
from openpyxl import Workbook
import json
import os

CONFIGURACAO = "config.json"

def carregar_configuracao():
    with open(CONFIGURACAO, "r", encoding="utf-8") as arquivo:
        config = json.load(arquivo)
    return config

def iniciar():
    config = carregar_configuracao()
    nome_planilha = config["planilha"]
    colunas = config["colunas"]

    if not os.path.exists(nome_planilha):
        wb = Workbook()
        ws = wb.active
        ws.title = "Lista de Compras"

        for col, nome in enumerate(colunas, start=1):
            ws.cell(row=1, column=col, value=nome)

        wb.save(nome_planilha)
        print(f"planilha '{nome_planilha}' criada com sucesso")
    else:
        print(f"planilha '{nome_planilha}' carregada com sucesso")

def _proximo_id(ws):
    ids = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            ids.append(row[0])
    return max(ids) + 1 if ids else 1

def adicionar(produto, quantidade):
    config = carregar_configuracao()
    nome_planilha = config["planilha"]

    try:
        wb = openpyxl.load_workbook(nome_planilha)
        ws = wb.active

        novo_id = _proximo_id(ws)
        status = "pendente"

        ws.append([novo_id, produto, quantidade, status])
        wb.save(nome_planilha)

        print(f"[PLANILHA] item adicionado!")
        print(f"  ID        : {novo_id}")
        print(f"  Produto   : {produto}")
        print(f"  Quantidade: {quantidade}")
        print(f"  Status    : {status}")

    except Exception as e:
        print(f"erro ao adicionar item: {e}")

def exibir():
    config = carregar_configuracao()
    nome_planilha = config["planilha"]

    try:
        wb = openpyxl.load_workbook(nome_planilha)
        ws = wb.active

        print("\n========== LISTA DE COMPRAS ==========")
        print(f"{'ID':<5} {'Produto':<20} {'Qtd':<15} {'Status':<10}")
        print("-" * 52)

        tem_itens = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                tem_itens = True
                print(f"{str(row[0]):<5} {str(row[1]):<20} {str(row[2]):<15} {str(row[3]):<10}")

        if not tem_itens:
            print("nenhum item cadastrado na lista")

        print("=" * 52)

    except Exception as e:
        print(f"erro ao exibir lista: {e}")

def remover(produto):
    config = carregar_configuracao()
    nome_planilha = config["planilha"]

    try:
        wb = openpyxl.load_workbook(nome_planilha)
        ws = wb.active

        removido = False
        for row in ws.iter_rows(min_row=2):
            if row[1].value and row[1].value.lower() == produto.lower():
                ws.delete_rows(row[0].row)
                removido = True
                print(f"[PLANILHA] '{produto}' removido com sucesso!")
                break

        if not removido:
            print(f"[PLANILHA] '{produto}' nao encontrado na lista")

        wb.save(nome_planilha)

    except Exception as e:
        print(f"erro ao remover item: {e}")

def organizar():
    config = carregar_configuracao()
    nome_planilha = config["planilha"]

    try:
        wb = openpyxl.load_workbook(nome_planilha)
        ws = wb.active

        dados = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                dados.append(list(row))

        # ordena por nome do produto em ordem alfabetica
        dados.sort(key=lambda x: str(x[1]).lower())

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.value = None

        for i, dado in enumerate(dados, start=1):
            dado[0] = i
            ws.append(dado)

        wb.save(nome_planilha)
        print("[PLANILHA] lista organizada em ordem alfabetica!")
        exibir()

    except Exception as e:
        print(f"erro ao organizar lista: {e}")

def atuar(acao, produto=None, quantidade=None, categoria=None):
    if acao == "adicionar":
        adicionar(produto, quantidade)
    elif acao == "exibir":
        exibir()
    elif acao == "remover":
        remover(produto)
    elif acao == "organizar":
        organizar()
    else:
        print(f"[PLANILHA] acao '{acao}' nao reconhecida")
